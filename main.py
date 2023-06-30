from pathlib import Path
import multiprocessing
from base64 import b64encode, b64decode
import argparse
import random
import time
from primitives import NIZK
from Crypto.PublicKey import ECC
from subroutines import deserialize_ep

from openpyxl import load_workbook, Workbook
from gmpy2 import powmod, invert, mul, f_mod, add
from texttable import Texttable
import threshold_crypto as tc

from curve import Curve
from parties import Voter, Teller
from util import (
    multi_dim_index,
    print_bb,
    find_entry_by_comm,
    calculate_voter_term,
)

# Note: The variable names used here may not reflect the actual names
# used in the paper since we have renamed them for PEP-8 compliance.
# However, some dictionary keys and messages still use names from the paper
# for brevity.


parser = argparse.ArgumentParser(
    description="Hyperion: Transparent End-to-End Verifiable Voting with Coercion Mitigation"
)
parser.add_argument(
    "voter_count", metavar="N", type=int, help="Number of voters"
)
parser.add_argument(
    "teller_count", metavar="T", type=int, help="Number of tellers"
)
parser.add_argument(
    "teller_threshold_count",
    metavar="K",
    type=int,
    help="Teller threshold value",
)
parser.add_argument(
    "-maxv",
    "--max-vote",
    metavar="MAX",
    type=str,
    help="Maximum vote value [Default: 1]",
)


args = parser.parse_args()

num_voters = 50
if (
    args.voter_count is not None
    and int(args.voter_count) > 0
    and int(args.voter_count) < 10000000
):
    num_voters = int(args.voter_count)
num_tellers = 5
if (
    args.teller_count is not None
    and int(args.teller_count) > 0
    and int(args.teller_count) < 100
):
    num_tellers = int(args.teller_count)
k = 3
if (
    args.teller_threshold_count is not None
    and int(args.teller_threshold_count) > 0
    and int(args.teller_threshold_count) < 100
):
    if int(args.teller_threshold_count) > num_tellers:
        print(
            "The teller threshold value must be less than the number of tellers."
        )
        exit()

vote_min = 0
vote_max = 2
if args.max_vote is not None and int(args.max_vote) > 1:
    vote_max = int(args.max_vote)

q1 = multiprocessing.Queue()
q2 = multiprocessing.Queue()

t_voting_single = 0
t_verification_single = 0
t_re_enc_mix_ver = 0
t_mixing = 0
t_decryption = 0

voters = []
tellers = []

bb = []
final_bb = []
verification_bb = []

teller_proofs = []

teller_sk = []
teller_public_key = ""
teller_registry = []

curve = Curve("P-256")


def poc_setup():
    """Sets up voter IDs and voter objects for 'vote_max' voters.
    Generates DSA key pairs for each voter.
    Picks a random vote value for each voter in the range
    ('vote_min':'vote_max').
    Adds all 'voter' objects to the 'voters' list.
    """
    for i in range(0, num_voters):
        id = "VT" + str(i)
        voter = Voter(curve, id, vote_min, vote_max)
        voter.generate_dsa_keys()
        voter.choose_vote_value()
        voters.append(voter)


def setup():
    """The setup phase of the protocol.
    Sets up 'num_tellers' teller objects.
    The teller public key and the threshold secret keys for
    'num_tellers' tally tellers are established.
    Adds all 'teller' objects to the 'tellers' list.
    """
    global teller_public_key
    global teller_sk
    teller_public_key, teller_sk = Teller.generate_threshold_keys(
        k, num_tellers, curve.get_pars()
    )
    for i in range(0, num_tellers):
        teller = Teller(curve, teller_sk[i], teller_public_key)
        tellers.append(teller)


def voting():
    """The voting phase of the protocol.
    For each 'voter' in the 'voters' list:
        a trapdoor keypair is generated,
        a proof of knowledge of the trapdoor secret key is generated,
        the vote is encrypted under the tellers' threshold public key,
        a proof of wellformedness of the ballot is generated,
        the signed, encrypted ballot is posted to a bulletin board.
    """
    for voter in voters:
        t_voting_single_start = time.time()
        voter.generate_trapdoor_keypair()
        voter.generate_pok_trapdoor_keypair()
        voter.encrypt_vote(teller_public_key)
        voter.generate_wellformedness_proof(teller_public_key)
        bb_data = voter.sign_ballot()
        bb.append(bb_data)
        t_voting = time.time() - t_voting_single_start
        global t_voting_single
        t_voting_single = t_voting_single + t_voting


def tallying():
    """The tallying phase of the protocol.
    For each 'ballot' on the bulletin board:
        the signature is verified,
        the proof of knowledge of the trapdoor secret key is verified,
        the proof of wellformedness of the ballot is verified,
    For each valid 'ballot' on the bulletin board:
        the tellers raise the public trapdoor key to
        a secret exponent 'r_i', and encrypt this value under
        their public key,
        the tellers produce a proof that the last step
        was performed correctly.
    The encrypted votes and 'h_r' tuples are subjected to a series of
    parallel re-encryption mixes by the tally tellers.
    The tuples are decrypted by the tally tellers and posted to
    a final bulletin board. The code in this phase has been modified to
    allow it to run faster on a multi-core system.
    """
    global final_bb
    global bb
    global t_mixing
    global t_decryption
    t_mixing_start = time.time()

    tagged_bb = []
    combined_bb = []
    raised_bb = []
    index = 0
    n = multiprocessing.cpu_count()
    for item in bb:
        temp = []
        temp.append(index)
        temp.append(item)
        tagged_bb.append(temp)
        index = index + 1
    k, m = divmod(len(tagged_bb), n)
    split_bb = [
        tagged_bb[i * k + min(i, m) : (i + 1) * k + min(i + 1, m)]
        for i in range(n)
    ]
    teller_proofs = []
    global teller_registry

    for teller in tellers:
        q1 = multiprocessing.Queue()
        q2 = multiprocessing.Queue()
        q3 = multiprocessing.Queue()

        processes = [
            multiprocessing.Process(
                target=teller.mp_raise_h, args=(ciph, q1, q2, q3)
            )
            for ciph in split_bb
        ]
        for p in processes:
            p.start()
        data_proofs = []
        data_registry = []
        data = []

        for p in processes:
            data_proofs = data_proofs + q1.get()
            data_registry = data_registry + q2.get()
            data = data + q3.get()

        for p in processes:
            p.join()
            # p.close()

        teller_proofs = teller_proofs + data_proofs
        teller_registry = teller_registry + data_registry
        data.sort()
        combined_bb.append(data)

    for i in range(len(combined_bb[0])):
        prod_a = deserialize_ep(combined_bb[0][i][1]["h_r"][0])
        prod_b = deserialize_ep(combined_bb[0][i][1]["h_r"][1])
        sum_r = combined_bb[0][i][1]["h_r"][2]
        for j in range(1, len(combined_bb)):
            prod_a = prod_a + deserialize_ep(combined_bb[j][i][1]["h_r"][0])
            prod_b = prod_b + deserialize_ep(combined_bb[j][i][1]["h_r"][1])
            sum_r = sum_r + (combined_bb[j][i][1]["h_r"][2])

        combined_bb[0][i][1]["h_r"] = (prod_a, prod_b, sum_r)
        temp = []
        temp.append(combined_bb[0][i][0])
        temp.append(combined_bb[0][i][1])
        raised_bb.append(temp)

    for i in range(0, len(raised_bb)):
        ballot = raised_bb[i][1]
        final_bb.append(
            {
                "ev": ballot["ev"],
                "h_r": ballot["h_r"],
                "proof_h_r": ballot["proof_h_r"],
            }
        )

    for i in range(0, len(teller_proofs)):
        record = teller_proofs[i]
        record["h_r"][0] = deserialize_ep(record["h_r"][0])
        record["h_r"][1] = deserialize_ep(record["h_r"][1])
        record["proof"]["t_1"] = deserialize_ep(record["proof"]["t_1"])
        record["proof"]["t_2"] = deserialize_ep(record["proof"]["t_2"])
        record["ptk"] = deserialize_ep(record["ptk"])
        """
        Teller.verify_proof_h_r(
            curve,
            teller_public_key,
            record["h_r"],
            record["ptk"],
            record["proof"],
            record["id"],
        )
        """

    list_0 = [[0] * 2] * len(final_bb)
    for i in range(0, len(final_bb)):
        temp = []
        temp.append(final_bb[i]["ev"])
        temp.append(final_bb[i]["h_r"])
        list_0[i] = temp

    previous = list_0
    global t_re_enc_mix_ver
    for i in range(1):
        teller = tellers[i]
        proof = teller.re_encryption_mix(previous)
        final_bb = proof[0]
        t_re_enc_mix_ver_start = time.time()
        # teller.verify_re_enc_mix(previous, proof)
        t_re_enc_mix_ver_end = time.time()
        t_re_enc_mix_ver = t_re_enc_mix_ver + (
            t_re_enc_mix_ver_end - t_re_enc_mix_ver_start
        )
        previous = final_bb
    t_mixing = (time.time() - t_mixing_start) - t_re_enc_mix_ver
    t_decryption_start = time.time()
    tagged_ciphertexts = teller.tag_ciphertexts(final_bb)
    split_ciphertexts = teller.ciphertext_list_split(
        tagged_ciphertexts, multiprocessing.cpu_count()
    )

    compound_pd = []
    compound_pd2 = []
    for teller in tellers:
        q1 = multiprocessing.Queue()
        q2 = multiprocessing.Queue()
        q3 = multiprocessing.Queue()
        processes = [
            multiprocessing.Process(
                target=teller.mp_partial_decrypt, args=(ciph, q1, q2, q3)
            )
            for ciph in split_ciphertexts
        ]
        for p in processes:
            p.daemon = True
            p.start()
        data = []
        data2 = []
        proofs = []
        for p in processes:
            data = data + q1.get()
            data2 = data2 + q2.get()
            proofs.append(q3.get())
        for p in processes:
            p.join()
            # p.close()
        compound_pd.append(data)
        compound_pd2.append(data2)

    final_pd = []
    final_pd2 = []

    for i in range(len(compound_pd[0])):
        temp = []
        temp.append(i)
        temp2 = []
        temp2.append(i)
        subtemp = []
        subtemp2 = []
        for item in compound_pd:
            sub_item = multi_dim_index(item, i)
            subtemp.append(sub_item[1])
        for item in compound_pd2:
            sub_item = multi_dim_index(item, i)
            subtemp2.append(sub_item[1])

        temp.append(subtemp)
        temp2.append(subtemp2)

        final_pd.append(temp)
        final_pd2.append(temp2)

    global decrypted
    split_ciphertexts = tellers[0].ciphertext_list_split(
        final_pd, multiprocessing.cpu_count()
    )
    processes = [
        multiprocessing.Process(
            target=tellers[0].mp_full_decrypt,
            args=(ciph, tagged_ciphertexts, 1, q1),
        )
        for ciph in split_ciphertexts
    ]
    for p in processes:
        p.daemon = True
        p.start()
    data = []
    for p in processes:
        data = data + q1.get()

    for p in processes:
        p.join()
        # p.close()
    vote_list = data
    split_ciphertexts = tellers[0].ciphertext_list_split(
        final_pd2, multiprocessing.cpu_count()
    )
    processes = [
        multiprocessing.Process(
            target=tellers[0].mp_full_decrypt,
            args=(ciph, tagged_ciphertexts, 2, q1),
        )
        for ciph in split_ciphertexts
    ]
    for p in processes:
        p.daemon = True
        p.start()
    data = []
    for p in processes:
        data = data + q1.get()

    for p in processes:
        p.join()
        # p.close()
    comm_list = data

    comm = None
    for item in vote_list:
        index = item[0]
        for subitem in comm_list:
            if subitem[0] == index:
                comm = subitem[1]
                break
        verification_bb.append({"v": item[1], "comm": comm})
    t_decryption = time.time() - t_decryption_start


def notification():
    """The tallying phase of the protocol.
    The 'r_i' term that corresponds to each voter is encrypted under
    their public key and sent privately to said voter.
    """
    for voter in voters:
        g_r = calculate_voter_term(curve, voter.id, teller_registry)
        voter.notify(g_r)


def verification():
    """The verification phase of the protocol.

    Every voter raises the value it receives from the tellers in the
    notification phase to its private trapdoor key. The voter then
    looks up the final bulletin board and checks if their vote has
    been recorded correctly.

    The program aborts if verification fails for any voter at
    this stage.
    """

    for i in range(len(voters)):
        voter = voters[i]
        t_verification_single_start = time.time()
        verification_comm = voter.generate_verification_comm()
        entry = find_entry_by_comm(verification_comm, verification_bb)
        if (
            ECC.EccPoint(entry["v"]["x"], entry["v"]["y"], entry["v"]["curve"])
            == voter.g_vote
        ):
            pass
        else:
            print("Error: Verification failed for voter" + str(voter.id))
            exit()
        t_verification = time.time() - t_verification_single_start
        global t_verification_single
        t_verification_single = t_verification_single + t_verification


def coercion_mitigation():
    """The coercion mitigation mechanism.

    A single voter selects another vote from the bulletin board and
    produces a fake dual key such that verification reveals a fake vote.
    """
    voter = voters[0]
    target = None
    # pick a random entry in the vbb
    for entry in verification_bb:
        if entry["v"] != voter.g_vote:
            target = entry
            break
    fake_dual_key = deserialize_ep(entry["comm"]) * (
        curve.get_pars().order - voter.secret_trapdoor_key
    )


def individual_views():
    """Individual Views.

    A unique view of a bulletin board is generated for a single voter.
    """
    voter = voters[random.randrange(0, len(voters))]
    iv = tellers[0].individual_board_shuffle(verification_bb)
    individual_view = iv[0]
    key = iv[1]
    g_rkey = calculate_voter_term(curve, voter.id, teller_registry) * key


def print_verification_bb():
    """Prints the contents of the final bulletin board to console."""
    table = Texttable()
    table.add_row(["Vote", "Commitment"])
    for item in verification_bb:
        comm_str = b64encode(str(hex(item['comm']['x'])).encode('UTF-8')).decode() 
        comm_str = comm_str + b64encode(str(hex(item['comm']['y'])).encode('UTF-8')).decode() 
        table.add_row([str(item["v"]), comm_str])
    print(table.draw())
    print()


print(
    "Hyperion: Transparent End-to-End Verifiable Voting with Coercion Mitigation"
)
print()

print("Running trials...")


poc_setup()


t_setup_start = time.time()
setup()
t_setup = str(time.time() - t_setup_start)

voting()
for i in range(0, len(bb)):
    ballot = bb[i]
    Teller.validate_ballot(curve, teller_public_key, ballot)

t_tallying_start = time.time()
tallying()
t_tallying = str((time.time() - t_tallying_start) - t_re_enc_mix_ver)

t_notification_start = time.time()
notification()
t_notification = str(time.time() - t_notification_start)

verification()

t_iv_start = time.time()
individual_views()
t_iv = str(time.time() - t_iv_start)


t_coercion_mitigation_start = time.time()
coercion_mitigation()
t_coercion_mitigation = str(time.time() - t_coercion_mitigation_start)

t_voting_single = str(t_voting_single / num_voters)
t_verification_single = str(t_verification_single / num_voters)

print_verification_bb()

print()
print("Voter count: " + str(num_voters))
print("Tally teller count: " + str(num_tellers))

table = Texttable()
output_headings = [
    "Setup",
    "Voting (avg.)",
    "Tallying (Mixing)",
    "Tallying (Decryption)",
    "Notification",
    "Verification (avg.)",
    "Coercion Mitigation",
    "Individual Views",
]

table.add_row(output_headings)
table.add_row(
    [
        t_setup,
        t_voting_single,
        t_mixing,
        t_decryption,
        t_notification,
        t_verification_single,
        t_coercion_mitigation,
        t_iv,
    ]
)


print(table.draw())

file_name = "Hyperion-Timing-Data.xlsx"
if not Path(file_name).exists():
    results_workbook = Workbook()
    results_counter = 1
    results_max_row = 0
else:
    results_workbook = load_workbook(file_name)
    results_max_row = results_workbook.active.max_row
    results_counter = (
        results_workbook.active["A" + str(results_max_row)].value + 1
    )
results_worksheet = results_workbook.active


if results_max_row == 0:
    results_worksheet.append(
        [
            "N",
            "Voters",
            "Tellers",
            "Threshold",
            "Setup",
            "Voting (avg.)",
            "Tallying (Mixing)",
            "Tallying (Decryption)",
            "Notification",
            "Verification (avg.)",
            "Coercion Mitigation",
            "Individual Views",
        ]
    )
timing_data = [
    results_counter,
    int(num_voters),
    int(num_tellers),
    int(k),
    float(t_setup),
    float(t_voting_single),
    float(t_mixing),
    float(t_decryption),
    float(t_notification),
    float(t_verification_single),
    float(t_coercion_mitigation),
    float(t_iv),
]
results_worksheet.append(timing_data)
results_workbook.save(file_name)

print(
    '\nThese values (all in seconds) have been written to a file named "Hyperion-Timing-Data.xlsx", in the current working directory.'
)
